"""
前晋四物流专线价表解析器
把 xlsx 的 9 个 sheet 解析成结构化 prices.json，供报价引擎与前端使用。

输出结构见 parse_workbook() 返回值。
"""
import re
import json
import openpyxl
from datetime import datetime, date

COUNTRY_SHEETS = {
    "US": "美国",
    "UK": "英国",
    "EU": "欧洲",
    "CA": "加拿大",
}
COUNTRY_NAMES = {"US": "美国", "UK": "英国", "EU": "欧洲", "CA": "加拿大"}

VOL_RE = re.compile(r"(\d+)\s*[Kk][Gg]\s*\+")
HEADER_NAME = "渠道名称"


def _norm_code(s):
    if s is None:
        return None
    return re.sub(r"\s+", "", str(s)).upper()


def _to_num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s in ("", "/", "*", "-", "—"):
        return None
    try:
        return float(s)
    except ValueError:
        m = re.search(r"(\d+(?:\.\d+)?)", s)
        return float(m.group(1)) if m else None


def _transport_class(transport):
    """根据运输方式推导报关费口径的渠道类型。"""
    if transport is None:
        return "专线"
    t = str(transport)
    if "UPS" in t or "快递" in t:
        return "快递"
    if "卡航" in t or "铁路" in t:
        return "卡航铁路"
    if "空运" in t or "空飞" in t or "直飞" in t or "转飞" in t:
        return "空运"
    if "海运" in t or "美森" in t or "船" in t:
        return "海运"
    return "专线"


def _vol_divisor(transport):
    """快递(UPS)按 /5000，其余专线按 /6000。"""
    return 5000 if _transport_class(transport) == "快递" else 6000


def _infer_class_from_name(name):
    """运输方式空缺时，按渠道名称关键词推断运输类别（用于报关费口径）。"""
    n = name or ""
    if "快递" in n or "UPS" in n:
        return "快递"
    if "卡航" in n or "铁路" in n:
        return "卡航铁路"
    if "空" in n:
        return "空运"
    if "海" in n:
        return "海运"
    return "专线"


def find_header_row(ws, max_scan=20):
    for r in range(1, max_scan + 1):
        if ws.cell(row=r, column=2).value == HEADER_NAME:
            return r
    return None


def parse_country(ws):
    """解析单个国家价表 sheet -> channels 列表。"""
    hdr = find_header_row(ws)
    if hdr is None:
        return []
    channels = []
    cur = None
    cur_tier_cols = []  # [(col_index, min_weight), ...]
    last_col = ws.max_column
    n = ws.max_row
    r = hdr

    def read_zone_fields(rr, inherit):
        transport = _cell_str(ws.cell(row=rr, column=9).value)
        return {
            "transport": transport or (inherit or {}).get("transport"),
            "pickup_lead": _cell_str(ws.cell(row=rr, column=10).value) or (inherit or {}).get("pickup_lead"),
            "sign_lead": _cell_str(ws.cell(row=rr, column=11).value) or (inherit or {}).get("sign_lead"),
            "comp_penalty": _cell_str(ws.cell(row=rr, column=12).value) or (inherit or {}).get("comp_penalty"),
            "remark": _cell_str(ws.cell(row=rr, column=13).value) or (inherit or {}).get("remark"),
        }

    while r <= n:
        name = ws.cell(row=r, column=2).value
        code = ws.cell(row=r, column=3).value
        if name == HEADER_NAME:  # 重新解析价格分层列
            cur_tier_cols = []
            for c in range(4, last_col + 1):
                hv = ws.cell(row=r, column=c).value
                if hv is None:
                    continue
                m = VOL_RE.search(str(hv))
                if m:
                    cur_tier_cols.append((c, int(m.group(1))))
            r += 1
            continue
        zone = ws.cell(row=r, column=4).value
        if name:  # 新渠道：本行自身即为第一个分区
            transport = _cell_str(ws.cell(row=r, column=9).value)
            tiers = {}
            for (c, mn) in cur_tier_cols:
                price = _to_num(ws.cell(row=r, column=c).value)
                if price is not None:
                    tiers[mn] = price
            f = read_zone_fields(r, None)
            cur = {
                "code": str(code).strip() if code else None,
                "name": str(name).strip(),
                "transport": f["transport"],
                "transport_class": _transport_class(transport) if transport else _infer_class_from_name(name),
                "vol_divisor": _vol_divisor(transport),
                "clearance_fee": _to_num(ws.cell(row=r, column=8).value) or 0,
                "pickup_lead": f["pickup_lead"],
                "sign_lead": f["sign_lead"],
                "comp_penalty": f["comp_penalty"],
                "remark": f["remark"],
                "zones": [],
                "tier_cols": [mn for (_, mn) in cur_tier_cols],
            }
            channels.append(cur)
            if zone:
                cur["zones"].append({
                    "zone": str(zone).strip(), "tiers": tiers,
                    "transport": f["transport"], "pickup_lead": f["pickup_lead"],
                    "sign_lead": f["sign_lead"], "comp_penalty": f["comp_penalty"],
                    "remark": f["remark"],
                })
        elif zone and cur is not None:  # 分区子行：继承渠道字段，价格取本行(若无则继承上一分区)
            tiers = {}
            has_price = False
            for (c, mn) in cur_tier_cols:
                price = _to_num(ws.cell(row=r, column=c).value)
                if price is not None:
                    tiers[mn] = price
                    has_price = True
            if not has_price and cur["zones"]:
                tiers = dict(cur["zones"][-1]["tiers"])
            f = read_zone_fields(r, cur)
            cur["zones"].append({
                "zone": str(zone).strip(), "tiers": tiers,
                "transport": f["transport"], "pickup_lead": f["pickup_lead"],
                "sign_lead": f["sign_lead"], "comp_penalty": f["comp_penalty"],
                "remark": f["remark"],
            })
        r += 1
    return channels


def _cell_str(v):
    if v is None:
        return None
    return str(v).strip()


def parse_fba(ws):
    """参考 sheet: 渠道/仓码/12KG+/75KG+/100KG+/时效 -> 仓码映射到渠道列表。

    同一个 FBA 仓码可能对应多个渠道（如 ONT8 同时属于海卡13天/21天），
    故用列表保留全部，解析与匹配时再按所选渠道或默认首个处理。
    """
    out = {}
    for r in range(2, ws.max_row + 1):
        channel = ws.cell(row=r, column=1).value
        code_raw = ws.cell(row=r, column=2).value
        if not channel or not code_raw:
            continue
        code = _norm_code(code_raw)
        tiers = {}
        for c, mn in [(3, 12), (4, 75), (5, 100)]:
            p = _to_num(ws.cell(row=r, column=c).value)
            if p is not None:
                tiers[mn] = p
        lead = _cell_str(ws.cell(row=r, column=6).value)
        entry = {"channel": str(channel).strip(), "tiers": tiers, "lead": lead}
        out.setdefault(code, []).append(entry)
    return out


def parse_vessel(ws):
    out = []
    for r in range(4, ws.max_row + 1):
        channel = ws.cell(row=r, column=2).value
        if not channel:
            continue
        def d(v):
            if isinstance(v, (datetime, date)):
                return v.strftime("%Y/%m/%d")
            return _cell_str(v)
        out.append({
            "channel": str(channel).strip().replace("\n", " / "),
            "carrier": _cell_str(ws.cell(row=r, column=3).value),
            "vessel": _cell_str(ws.cell(row=r, column=4).value),
            "etd": d(ws.cell(row=r, column=5).value),
            "eta": d(ws.cell(row=r, column=6).value),
            "port": _cell_str(ws.cell(row=r, column=7).value),
            "cut_off": _cell_str(ws.cell(row=r, column=8).value),
        })
    return out


def parse_notice(ws):
    """发货须知 sheet -> 结构化规则。"""
    notice = {
        "currency": "RMB",
        "volumetric": {"快递": 5000, "专线": 6000},
        "girth_formula": None,
        "piece_girth_range": None,
        "piece_weight_range": None,
        "oversize_fee": {},
        "overweight_fee": {},
        "export_customs_fee": {},   # 出口报关费, by 渠道类型 -> {国家:费用}
        "import_clears_fee": {},    # 进口清关费, by 渠道类型 -> {国家:费用}
        "rules": [],                # 其余规则(文本展示)
        "long_text": [],             # 长文本段落(免赔/赔偿/索赔等)
    }
    last_node = None  # 发货须知「专线节点」列为合并单元格，仅首行有值，向后继承
    last_charge = None  # 「收费类型」列同样可能合并，向后继承
    long_text_nodes = ("时效免赔说明", "签收免赔情况说明", "赔偿标准", "索赔程序", "保险索赔流程")
    for r in range(3, ws.max_row + 1):
        raw_node = _cell_str(ws.cell(row=r, column=2).value)
        node = raw_node if raw_node else last_node
        if raw_node:
            last_node = raw_node
        raw_charge = _cell_str(ws.cell(row=r, column=3).value)
        charge = raw_charge if raw_charge else last_charge
        if raw_charge and node not in long_text_nodes:
            last_charge = raw_charge
        ctype = _cell_str(ws.cell(row=r, column=4).value)
        us = _cell_str(ws.cell(row=r, column=5).value)
        uk = _cell_str(ws.cell(row=r, column=6).value)
        eu = _cell_str(ws.cell(row=r, column=7).value)
        ca = _cell_str(ws.cell(row=r, column=8).value)
        note = _cell_str(ws.cell(row=r, column=9).value)
        if not node:
            continue
        # 长文本段落（免赔/赔偿/索赔/保险），文本在 col3
        if node in long_text_nodes:
            notice["long_text"].append({
                "title": node,
                "text": (charge or "").replace("\\n", "\n").strip(),
                "note": note,
            })
            continue
        by_country = {}
        for k, v in [("US", us), ("UK", uk), ("EU", eu), ("CA", ca)]:
            if v not in (None, "", "/"):
                by_country[k] = v
        # 结构化数值规则
        if node == "计费重" and charge == "材积重公式":
            pass  # already in volumetric
        elif node == "计费重" and charge == "围长公式":
            notice["girth_formula"] = us
        elif node == "计费重" and charge == "单件围长范围":
            notice["piece_girth_range"] = f"{us}；{note or ''}".strip("；")
        elif node == "计费重" and charge == "单件实重范围":
            notice["piece_weight_range"] = f"{us}；{note or ''}".strip("；")
        elif node == "计费重" and charge == "超长收费":
            notice["oversize_fee"] = {k: _to_num(v) for k, v in by_country.items() if _to_num(v) is not None}
        elif node == "计费重" and charge == "超重收费":
            notice["overweight_fee"] = {k: _to_num(v) for k, v in by_country.items() if _to_num(v) is not None}
        elif node == "出口报关" and charge == "报关费":
            notice["export_customs_fee"][ctype] = {k: _to_num(v) for k, v in by_country.items() if _to_num(v) is not None}
        elif node == "进口清关" and charge == "清关费":
            notice["import_clears_fee"][ctype] = {k: _to_num(v) for k, v in by_country.items() if _to_num(v) is not None}
        else:
            notice["rules"].append({
                "node": node, "charge": charge, "ctype": ctype,
                "by_country": by_country, "note": note,
            })
    return notice


def parse_directory(ws):
    warehouses = []
    for r in range(5, 8):
        v = ws.cell(row=r, column=2).value
        if v:
            s = str(v).strip()
            # 拆成 {name, addr} 结构化对象（格式："仓名：地址 联系人 电话 时间"）
            idx = s.find("：")
            if idx > 0:
                warehouses.append({"name": s[:idx].strip(), "addr": s[idx + 1:].strip()})
            else:
                warehouses.append({"name": s, "addr": ""})
    website = None
    for r in range(1, 10):
        v = ws.cell(row=r, column=2).value
        if v and "aheadfour" in str(v).lower():
            website = str(v).strip()
            break
    return {"warehouses": warehouses, "website": website}


def parse_workbook(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    countries = {}
    for key, sheet in COUNTRY_SHEETS.items():
        ws = wb[sheet]
        countries[key] = {
            "name": COUNTRY_NAMES[key],
            "channels": parse_country(ws),
        }
    # FBA 映射：把参考 sheet 的渠道名关联到国家渠道代码
    fba = parse_fba(wb["参考"])
    # 建立 渠道名->code 映射（从国家渠道）
    name_to_code = {}
    for key, cobj in countries.items():
        for ch in cobj["channels"]:
            name_to_code[ch["name"]] = ch["code"]
    for code, entries in fba.items():
        for entry in entries:
            entry["code"] = name_to_code.get(entry["channel"])
    data = {
        "version": None,
        "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "source_file": None,
        "company": parse_directory(wb["目录"]),
        "countries": countries,
        "fba_map": fba,
        "shipping_notice": parse_notice(wb["发货须知"]),
        "vessel_schedule": parse_vessel(wb["船期表"]),
    }
    return data


if __name__ == "__main__":
    import sys
    src = sys.argv[1] if len(sys.argv) > 1 else r"D:/Documents/WXWork/1688856495771770/Cache/File/2026-08/前晋四物流专线价表20260803.xlsx"
    out = parse_workbook(src)
    # 校验摘要
    print("countries:", {k: len(v["channels"]) for k, v in out["countries"].items()})
    print("fba entries:", len(out["fba_map"]))
    print("oversize_fee:", out["shipping_notice"]["oversize_fee"])
    print("overweight_fee:", out["shipping_notice"]["overweight_fee"])
    print("export_customs_fee keys:", list(out["shipping_notice"]["export_customs_fee"].keys()))
    print("rules count:", len(out["shipping_notice"]["rules"]))
    print("long_text count:", len(out["shipping_notice"]["long_text"]))
    print("vessel rows:", len(out["vessel_schedule"]))
    # 抽样一个渠道
    us = out["countries"]["US"]["channels"][0]
    print("sample channel:", us["name"], us["code"], "zones:", len(us["zones"]), "tiers:", us["tier_cols"])
    # 落盘
    with open(sys.argv[2] if len(sys.argv) > 2 else "prices_test.json", "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)
    print("written.")
